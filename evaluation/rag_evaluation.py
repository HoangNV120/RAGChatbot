"""
RAG Chat Evaluation Script

This script evaluates the performance of the RAG chat system by:
1. Loading test questions and ground truth answers from an Excel file
2. Running each question through the RAG chat system
3. Calculating evaluation metrics (latency, etc.)
4. Saving results to an Excel file with detailed results and summary statistics
"""

import asyncio
import pandas as pd
import numpy as np
import time
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import traceback

# Add the project root to the Python path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Import RAG components
from app.rag_chat import RAGChat
from app.vector_store import VectorStore


class RAGEvaluator:
    def __init__(self):
        """Initialize the RAG evaluator"""
        self.rag_chat = None
        
    async def setup_rag_components(self):
        """Set up the RAG components for evaluation"""
        print("🔧 Setting up RAG components...")
        # Initialize vector store
        vector_store = VectorStore()
        # Initialize RAG chat with the vector store
        self.rag_chat = RAGChat(vector_store=vector_store)
        print("✅ RAG components initialized")

    def load_test_data(self, excel_path: str) -> pd.DataFrame:
        """
        Load test data from Excel file
        
        Args:
            excel_path: Path to the Excel file containing test data
            
        Returns:
            DataFrame with questions and ground truth answers
        """
        try:
            print(f"📊 Loading test data from {excel_path}...")
            df = pd.read_excel(excel_path)
            
            # Check if required columns exist
            required_columns = ['question', 'answer']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                raise ValueError(f"Missing required columns in Excel file: {', '.join(missing_columns)}")
            
            print(f"✅ Loaded {len(df)} test questions")
            return df
            
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            raise

    async def evaluate_question(self, question: str):
        """
        Evaluate a single question using the RAG chat system
        
        Args:
            question: The question to evaluate
            
        Returns:
            Dictionary with response and metrics
        """
        try:
            start_time = time.time()
            
            # Get response from RAG chat
            response = await self.rag_chat.generate_response(question)
            
            end_time = time.time()
            latency = end_time - start_time
            
            # Extract the actual answer text from the response
            answer_text = response.get("output", "")
            
            # Get contexts used for the response
            contexts = []
            if "docs" in response and response["docs"]:
                for doc in response["docs"]:
                    if hasattr(doc, "page_content"):
                        contexts.append(doc.page_content)
            
            # Get model used if available
            model_used = response.get("model", "unknown")
            
            return {
                "chatbot_response": answer_text,
                "latency": latency,
                "contexts": contexts,
                "model_used": model_used
            }
            
        except Exception as e:
            print(f"❌ Error evaluating question: {e}")
            return {
                "chatbot_response": f"Error: {str(e)}",
                "latency": -1,
                "contexts": [],
                "model_used": "error"
            }

    async def run_evaluation(self, input_excel: str, output_excel: str):
        """
        Run the full evaluation process
        
        Args:
            input_excel: Path to the input Excel file with test data
            output_excel: Path to save the evaluation results
        """
        try:
            # Setup RAG components
            await self.setup_rag_components()
            
            # Load test data
            df = self.load_test_data(input_excel)
            
            print(f"🚀 Starting evaluation of {len(df)} questions...")
            
            # Create results dataframe
            results = []
            
            # Process each question
            for idx, row in df.iterrows():
                question = row['question']
                ground_truth = row['answer']
                
                print(f"[{idx+1}/{len(df)}] Evaluating: {question[:50]}...")
                
                # Evaluate the question
                eval_result = await self.evaluate_question(question)
                
                # Add to results
                result = {
                    "question": question,
                    "ground_truth": ground_truth,
                    "chatbot_response": eval_result["chatbot_response"],
                    "latency": eval_result["latency"],
                    "model_used": eval_result["model_used"]
                }
                
                results.append(result)
                
                # Print progress
                if eval_result["latency"] > 0:
                    print(f"  ✓ Latency: {eval_result['latency']:.2f}s, Model: {eval_result['model_used']}")
                else:
                    print(f"  ✗ Error occurred")
            
            # Convert results to DataFrame
            results_df = pd.DataFrame(results)
            
            # Save results to Excel
            self.save_results_to_excel(results_df, output_excel)
            
            print(f"✅ Evaluation completed! Results saved to: {output_excel}")
            
            return results_df
            
        except Exception as e:
            print(f"❌ Error in evaluation process: {e}")
            traceback.print_exc()
            raise

    def save_results_to_excel(self, results_df: pd.DataFrame, output_path: str):
        """
        Save evaluation results to Excel with formatting
        
        Args:
            results_df: DataFrame with evaluation results
            output_path: Path to save the Excel file
        """
        try:
            print(f"💾 Saving results to {output_path}...")
            
            # Create a new workbook
            wb = Workbook()
            
            # Create detailed results sheet
            detailed_sheet = wb.active
            detailed_sheet.title = "Detailed Results"
            
            # Add data to detailed sheet
            for r in dataframe_to_rows(results_df, index=False, header=True):
                detailed_sheet.append(r)
            
            # Format header row
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in detailed_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust column widths
            for column in detailed_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, min(cell_length, 50))  # Cap at 50 chars
                
                adjusted_width = max_length + 2
                detailed_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create summary sheet
            summary_sheet = wb.create_sheet(title="Summary")
            
            # Calculate summary statistics
            summary_data = {
                "Metric": ["Total Questions", "Average Latency (s)", "Min Latency (s)", 
                           "Max Latency (s)", "Median Latency (s)"],
                "Value": [
                    len(results_df),
                    results_df["latency"].mean(),
                    results_df["latency"].min(),
                    results_df["latency"].max(),
                    results_df["latency"].median()
                ]
            }
            
            # Add model usage statistics if available
            if "model_used" in results_df.columns:
                model_counts = results_df["model_used"].value_counts()
                for model, count in model_counts.items():
                    summary_data["Metric"].append(f"Model: {model}")
                    summary_data["Value"].append(f"{count} ({count/len(results_df)*100:.1f}%)")
            
            # Add summary data to sheet
            summary_df = pd.DataFrame(summary_data)
            for r in dataframe_to_rows(summary_df, index=False, header=True):
                summary_sheet.append(r)
            
            # Format summary sheet
            for cell in summary_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths for summary
            for column in summary_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                
                adjusted_width = max_length + 2
                summary_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            wb.save(output_path)
            
        except Exception as e:
            print(f"❌ Error saving results to Excel: {e}")
            raise


async def main():
    """Main function to run the evaluation"""
    try:
        # Define fixed paths in the evaluation folder
        current_dir = os.path.dirname(os.path.abspath(__file__))
        input_excel = os.path.join(current_dir, "data_test.xlsx")
        
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_excel = os.path.join(current_dir, f"evaluation_results_{timestamp}.xlsx")
        
        print(f"Input file: {input_excel}")
        print(f"Output will be saved to: {output_excel}")
        
        # Create evaluator and run evaluation
        evaluator = RAGEvaluator()
        await evaluator.run_evaluation(input_excel, output_excel)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        traceback.print_exc()


if __name__ == "__main__":
    asyncio.run(main()) 